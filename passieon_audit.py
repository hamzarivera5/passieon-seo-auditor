"""
Passieon SEO Auditor v3.0
Unlimited Crawl + Ahrefs + AI Strategic Analysis (qwen2.5:32b)
Generates 5-tab client-ready audit matching professional format.
"""

import customtkinter as ctk
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import time, re, json, os, sys, threading
from datetime import datetime, date
from collections import Counter, defaultdict

CRAWL_DELAY = 0.3
REQUEST_TIMEOUT = 15
AHREFS_BASE = "https://api.ahrefs.com/v3"
OLLAMA_MODEL = "qwen2.5:32b"
CONFIG_FILE = "passieon_config.json"
HEADERS = {"User-Agent": "PassieonSEOAuditor/3.0", "Accept": "text/html,application/xhtml+xml"}
COUNTRIES = [
    "Global (all countries)", "US - United States", "GB - United Kingdom", "PH - Philippines",
    "AU - Australia", "CA - Canada", "DE - Germany", "FR - France", "IN - India",
    "ID - Indonesia", "JP - Japan", "MY - Malaysia", "SG - Singapore", "AE - UAE",
    "SA - Saudi Arabia", "ES - Spain", "IT - Italy", "NL - Netherlands", "BR - Brazil",
]

def get_config_path():
    if getattr(sys, 'frozen', False): return os.path.join(os.path.dirname(sys.executable), CONFIG_FILE)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), CONFIG_FILE)
def load_config():
    p = get_config_path()
    if os.path.exists(p):
        try:
            with open(p, "r") as f: return json.load(f)
        except: pass
    return {}
def save_config(cfg):
    with open(get_config_path(), "w") as f: json.dump(cfg, f, indent=2)

# ─── OLLAMA ──────────────────────────────────────────────────────────────────
def check_ollama():
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=5)
        if r.status_code == 200:
            return any(OLLAMA_MODEL in m["name"] for m in r.json().get("models", []))
    except: pass
    return False

def ask_ollama(prompt, temperature=0.3, max_tokens=4000):
    try:
        r = requests.post("http://localhost:11434/api/generate",
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False,
                  "options": {"temperature": temperature, "num_predict": max_tokens, "num_ctx": 8192}}, timeout=600)
        if r.status_code == 200: return r.json().get("response", "").strip()
    except Exception as e: print(f"Ollama error: {e}")
    return ""

# ─── AHREFS ──────────────────────────────────────────────────────────────────
class AhrefsAPI:
    def __init__(self, api_key):
        self.headers = {"Authorization": f"Bearer {api_key}", "Accept": "application/json"}
        self.today = date.today().strftime("%Y-%m-%d")
    def _get(self, endpoint, params):
        params.setdefault("date", self.today); params.setdefault("output", "json")
        try:
            r = requests.get(f"{AHREFS_BASE}/{endpoint}", headers=self.headers, params=params, timeout=30)
            r.raise_for_status(); return r.json()
        except Exception as e:
            print(f"Ahrefs error [{endpoint}]: {e}")
            return None
    def domain_rating(self, domain):
        d = self._get("site-explorer/domain-rating", {"target": domain})
        return d.get("domain_rating", {}) if d else {}
    def overview_metrics(self, domain, country=None):
        p = {"target": domain, "mode": "subdomains"}
        if country: p["country"] = country
        d = self._get("site-explorer/metrics", p); return d.get("metrics", {}) if d else {}
    def backlinks_stats(self, domain):
        d = self._get("site-explorer/backlinks-stats", {"target": domain, "mode": "subdomains"})
        return d.get("metrics", {}) if d else {}
    def organic_keywords(self, domain, country=None, limit=200):
        p = {"target": domain, "mode": "subdomains",
             "select": "keyword,volume,best_position,sum_traffic,best_position_url,keyword_difficulty,cpc",
             "order_by": "sum_traffic:desc", "limit": limit}
        if country: p["country"] = country
        d = self._get("site-explorer/organic-keywords", p); return d.get("keywords", []) if d else []
    def top_pages(self, domain, limit=100):
        p = {"target": domain, "mode": "subdomains",
             "select": "url,sum_traffic,keywords,referring_domains,top_keyword,top_keyword_best_position,top_keyword_volume,value",
             "order_by": "sum_traffic:desc", "limit": limit}
        d = self._get("site-explorer/top-pages", p); return d.get("pages", []) if d else []
    def referring_domains(self, domain, limit=100):
        p = {"target": domain, "mode": "subdomains",
             "select": "domain_rating,domain,backlinks,first_seen,last_visited",
             "order_by": "domain_rating:desc", "limit": limit}
        d = self._get("site-explorer/referring-domains", p); return d.get("refdomains", []) if d else []
    def organic_competitors(self, domain, limit=10):
        p = {"target": domain, "mode": "subdomains",
             "select": "domain,org_keywords,org_traffic,common_keywords,keywords_unique",
             "order_by": "common_keywords:desc", "limit": limit}
        d = self._get("site-explorer/organic-competitors", p); return d.get("competitors", []) if d else []

# ─── CRAWLER ─────────────────────────────────────────────────────────────────
def normalize_url(url):
    if not url.startswith(("http://", "https://")): url = "https://" + url
    return url.rstrip("/")

def extract_links(soup, base_url, domain):
    links = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("#", "mailto:", "tel:", "javascript:")): continue
        full = urljoin(base_url, href).split("#")[0].split("?")[0].rstrip("/")
        if urlparse(full).netloc == domain: links.add(full)
    return links

def analyze_page(url, response, soup):
    issues = []; domain = urlparse(url).netloc
    data = {"url": url, "status_code": response.status_code, "title": "", "title_length": 0,
            "meta_description": "", "meta_desc_length": 0, "h1_count": 0, "h1_text": "",
            "h2_count": 0, "h2_texts": [], "word_count": 0, "image_count": 0, "images_missing_alt": 0,
            "internal_links": 0, "external_links": 0, "has_canonical": False,
            "has_og_tags": False, "has_schema": False, "schema_types": [],
            "load_time": 0, "issues": [], "redirect_chain": False}

    if response.status_code != 200: issues.append(f"Non-200 status: {response.status_code}")
    data["load_time"] = round(response.elapsed.total_seconds(), 2)
    if data["load_time"] > 3: issues.append("Slow page (>3s)")
    if len(response.history) > 0: data["redirect_chain"] = True; issues.append(f"Redirect chain ({len(response.history)} hops)")

    title_tag = soup.find("title")
    if title_tag and title_tag.string:
        data["title"] = title_tag.string.strip(); data["title_length"] = len(data["title"])
        if data["title_length"] < 30: issues.append("Title too short (<30 chars)")
        elif data["title_length"] > 60: issues.append("Title over 60 characters")
    else: issues.append("Missing title tag")

    meta_desc = soup.find("meta", attrs={"name": re.compile(r"description", re.I)})
    if meta_desc and meta_desc.get("content"):
        data["meta_description"] = meta_desc["content"].strip(); data["meta_desc_length"] = len(data["meta_description"])
        if data["meta_desc_length"] < 70: issues.append("Meta description below 70 characters")
        elif data["meta_desc_length"] > 155: issues.append("Meta description over 155 characters")
    else: issues.append("Missing meta description")

    h1s = soup.find_all("h1"); data["h1_count"] = len(h1s)
    if h1s: data["h1_text"] = h1s[0].get_text(strip=True)[:120]
    if len(h1s) == 0: issues.append("H1 tag missing")
    elif len(h1s) > 1: issues.append(f"Multiple H1 tags ({len(h1s)})")

    h2s = soup.find_all("h2"); data["h2_count"] = len(h2s)
    data["h2_texts"] = [h.get_text(strip=True)[:80] for h in h2s[:10]]
    h2_texts_lower = [h.get_text(strip=True).lower() for h in h2s]
    if len(h2_texts_lower) != len(set(h2_texts_lower)) and len(h2s) > 1: issues.append("Duplicate H2 tags")

    body = soup.find("body")
    if body:
        data["word_count"] = len(body.get_text(separator=" ", strip=True).split())
        if data["word_count"] < 300: issues.append("Thin content (<300 words)")

    imgs = soup.find_all("img"); data["image_count"] = len(imgs)
    ma = sum(1 for img in imgs if not img.get("alt")); data["images_missing_alt"] = ma
    if ma > 0: issues.append(f"{ma} images missing alt text")

    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith(("#", "mailto:", "tel:", "javascript:")): continue
        full = urljoin(url, href)
        if urlparse(full).netloc == domain: data["internal_links"] += 1
        else: data["external_links"] += 1

    canonical = soup.find("link", rel="canonical"); data["has_canonical"] = canonical is not None
    if not canonical: issues.append("Missing canonical tag")
    og = soup.find("meta", property=re.compile(r"og:", re.I)); data["has_og_tags"] = og is not None
    if not og: issues.append("Missing Open Graph tags")
    schemas = soup.find_all("script", type="application/ld+json"); data["has_schema"] = len(schemas) > 0
    for s in schemas:
        try:
            sd = json.loads(s.string)
            if isinstance(sd, dict): data["schema_types"].append(sd.get("@type", "Unknown"))
            elif isinstance(sd, list):
                for item in sd:
                    if isinstance(item, dict): data["schema_types"].append(item.get("@type", "Unknown"))
        except: pass
    if not data["has_schema"]: issues.append("No structured data (schema.org)")

    data["issues"] = issues
    return data

# ─── ISSUE GROUPING ─────────────────────────────────────────────────────────
ISSUE_META = {
    "Missing title tag": ("Critical", "Every page needs a unique title tag with primary keyword (50-60 chars).", "Pages without titles may not rank or show the URL in SERPs."),
    "Title too short (<30 chars)": ("Medium", "Expand titles to 50-60 characters including primary keyword and brand.", "Short titles waste SERP real estate and miss keyword targeting."),
    "Title over 60 characters": ("High", "Shorten titles to under 60 characters. Consider shortening brand suffix.", "Google truncates titles over ~60 characters."),
    "Missing meta description": ("Critical", "Write unique meta descriptions (120-155 chars) with target keyword and CTA.", "Google auto-generates snippets which are often poor and reduce CTR."),
    "Meta description below 70 characters": ("Medium", "Expand meta descriptions to 120-155 characters with benefits and keywords.", "Short descriptions waste SERP real estate."),
    "Meta description over 155 characters": ("High", "Trim meta descriptions to 120-155 characters. Front-load key info.", "Google truncates descriptions over ~155 characters."),
    "H1 tag missing": ("Critical", "Add a single H1 tag containing the primary keyword.", "H1 is the most important heading signal for page topic."),
    "Multiple H1 tags": ("Medium", "Keep one H1 per page. Change extras to H2 or H3.", "Multiple H1s confuse heading hierarchy."),
    "Duplicate H2 tags": ("Medium", "Remove or rename duplicate H2 headings.", "Duplicate headings confuse hierarchy and accessibility."),
    "Thin content (<300 words)": ("High", "Expand content to 600+ words with unique, valuable information.", "Thin pages have insufficient content to rank."),
    "Slow page (>3s)": ("High", "Optimize images, enable caching, minimize JS/CSS, consider CDN.", "Google uses page speed as a ranking factor."),
    "Images missing alt text": ("Critical", "Add descriptive alt text to all images.", "Missing alt = no Google Image indexing + accessibility violation."),
    "Missing canonical tag": ("Medium", "Add rel=canonical to prevent duplicate content issues.", "Without canonical tags, Google may index duplicates."),
    "Missing Open Graph tags": ("Low", "Add og:title, og:description, og:image for social sharing.", "Social shares show generic previews."),
    "No structured data (schema.org)": ("Critical", "Add JSON-LD: Organization globally, Product on products, Article on blog.", "Missing rich results in SERPs."),
    "Redirect chain": ("High", "Update internal links to point directly to final URL.", "Each redirect wastes crawl budget and adds latency."),
    "Non-200 status": ("Critical", "Fix or redirect pages returning error codes.", "Error pages cannot be indexed."),
}

def group_issues(pages_data):
    grouped = defaultdict(lambda: {"pages": [], "details": {}})
    for pg in pages_data:
        for iss in pg["issues"]:
            key = iss
            for pattern in ISSUE_META:
                if iss.startswith(pattern.split("(")[0].split(":")[0].strip()[:20]):
                    key = pattern; break
            entry = grouped[key]
            entry["pages"].append(pg["url"])
            path = urlparse(pg["url"]).path or "/"
            if "alt text" in key: entry["details"][pg["url"]] = f"{pg['images_missing_alt']} images"
            elif "Thin" in key: entry["details"][pg["url"]] = f"{pg['word_count']} words"
            elif "Slow" in key: entry["details"][pg["url"]] = f"{pg['load_time']}s"
            elif "over 60" in key: entry["details"][pg["url"]] = f'{pg["title_length"]} chars: "{pg["title"][:55]}..."'
            elif "over 155" in key: entry["details"][pg["url"]] = f"{pg['meta_desc_length']} chars"
            elif "below 70" in key: entry["details"][pg["url"]] = f'{pg["meta_desc_length"]} chars: "{pg["meta_description"][:40]}"'
            elif "Redirect" in key: entry["details"][pg["url"]] = f"{len([h for h in []])} hops"
    return grouped

# ─── BUILD DATA SUMMARY FOR AI ──────────────────────────────────────────────
def build_full_summary(domain, pages_data, ahrefs_data):
    dr = ahrefs_data.get("domain_rating", {}); met = ahrefs_data.get("metrics", {})
    bl = ahrefs_data.get("backlinks_stats", {}); kws = ahrefs_data.get("organic_keywords", [])
    tpg = ahrefs_data.get("top_pages", []); rds = ahrefs_data.get("referring_domains", [])
    comps = ahrefs_data.get("competitors", [])

    kw_lines = "\n".join(f'  "{k.get("keyword","")}" | pos:{k.get("best_position","?")} | vol:{k.get("volume",0)} | KD:{k.get("keyword_difficulty","?")} | traffic:{k.get("sum_traffic",0)} | url:{k.get("best_position_url","")}' for k in kws[:30])
    pg_lines = "\n".join(f'  {p.get("url","")} | traffic:{p.get("sum_traffic",0)} | kws:{p.get("keywords",0)} | refdoms:{p.get("referring_domains",0)} | top_kw:"{p.get("top_keyword","")}"' for p in tpg[:20])
    rd_lines = "\n".join(f'  {r.get("domain","")} | DR:{r.get("domain_rating",0)} | links:{r.get("backlinks",0)}' for r in rds[:30])
    comp_lines = "\n".join(f'  {c.get("domain","")} | kws:{c.get("org_keywords",0)} | traffic:{c.get("org_traffic",0)} | common:{c.get("common_keywords",0)}' for c in comps[:5])

    # Crawl stats
    grouped = group_issues(pages_data)
    issue_lines = "\n".join(f'  {iss}: {len(d["pages"])} pages' for iss, d in sorted(grouped.items(), key=lambda x: len(x[1]["pages"]), reverse=True))
    schema_found = set()
    for p in pages_data:
        for st in p.get("schema_types", []): schema_found.add(st)

    return f"""DOMAIN: {domain}
DR: {dr.get('domain_rating','N/A')} | Rank: {dr.get('ahrefs_rank','N/A')}
Organic Traffic: {met.get('org_traffic',0):,}/mo | Keywords: {met.get('org_keywords',0):,} | Top3: {met.get('org_keywords_1_3',0):,}
Traffic Value: ${met.get('org_cost',0)/100:,.0f}/mo
Live Backlinks: {bl.get('live',0):,} | Referring Domains: {bl.get('live_refdomains',0):,}
Pages Crawled: {len(pages_data)}
Schema types found: {', '.join(schema_found) if schema_found else 'None'}

ORGANIC COMPETITORS:
{comp_lines if comp_lines else '  None found'}

TOP KEYWORDS:
{kw_lines if kw_lines else '  None'}

TOP PAGES:
{pg_lines if pg_lines else '  None'}

REFERRING DOMAINS:
{rd_lines if rd_lines else '  None'}

CRAWL ISSUES:
{issue_lines if issue_lines else '  None'}"""

# ─── AI STRATEGIC ANALYSIS ──────────────────────────────────────────────────
def ai_content_gap(summary, domain, country_name):
    prompt = f"""You are a senior SEO strategist writing a Content Gap Analysis for {domain}.

Based on the audit data below, identify 15-20 keyword opportunities that this site SHOULD be targeting but isn't ranking well for (position >10 or not ranking). Group them into two sections:

SECTION 1: "PRODUCT & SERVICE KEYWORDS" (existing pages can target these)
SECTION 2: "HIGH-VOLUME CONTENT KEYWORDS" (need new blog posts)

For each keyword, output ONE line in this exact pipe-delimited format:
SECTION|keyword|volume|KD|CPC|currently_ranking|page_to_target|recommendation

Rules:
- Use actual keywords from the data OR suggest obvious related keywords based on what the site sells/does
- "currently_ranking" = position if ranking, or "Not ranking"
- "page_to_target" = existing URL or "New blog post"
- "recommendation" must be specific: include a suggested blog title if new content is needed
- Prioritize KD 0-10 keywords (easy wins)
- No headers, no explanations, just the pipe-delimited rows

{summary}

OUTPUT:"""
    return ask_ollama(prompt, max_tokens=4000)

def ai_backlink_strategy(summary, domain):
    prompt = f"""You are a senior SEO link building strategist writing a Backlink Strategy for {domain}.

Based on the referring domains and backlink data below, produce TWO sections:

SECTION 1: "CURRENT BACKLINK PROFILE" - Analyze each referring domain listed. For each, output:
domain|DR|links|type_assessment|quality|action|notes

Where:
- type_assessment = what kind of link (e.g., "Parent brand dofollow", "Event listing", "SPAM: link seller")
- quality = "Legitimate", "Needs audit", or "Toxic"
- action = "Keep", "Audit individually", or "DISAVOW"

SECTION 2: "LINK BUILDING TACTICS" - Suggest 10-15 specific link building tactics. For each:
tactic|priority|DR_range|effort|approach|expected_outcome|timeline

Rules:
- Be specific to this domain's industry/niche
- Include: citations, guest posts, link inserts, digital PR, partner links, unlinked mentions
- Reference actual competitor domains or industry sites where possible
- No generic advice. Every row must be actionable.
- No headers or explanations, just the pipe-delimited rows
- Separate sections with a line that says "---SECTION2---"

{summary}

OUTPUT:"""
    return ask_ollama(prompt, max_tokens=5000)

def ai_blog_strategy(summary, domain):
    prompt = f"""You are an SEO content strategist writing a Blog Content Strategy for {domain}.

Based on the keyword data and content gaps, recommend 10-15 specific blog posts to create. For each post, output ONE line:
priority|blog_title|target_keyword|volume|KD|word_count|internal_link_to|content_notes

Rules:
- Priority = P1 (publish first, highest impact), P2, or P3
- blog_title must be a specific, SEO-optimized title (not generic)
- target_keyword = primary keyword to rank for
- internal_link_to = specific existing URL on the site to link to (from the top pages data)
- content_notes = what to include (e.g., "Include before/after photos, day-by-day recovery timeline")
- Order by priority then by search volume
- Focus on KD 0-10 keywords first (quick wins)
- No headers or explanations, just the pipe-delimited rows

{summary}

OUTPUT:"""
    return ask_ollama(prompt, max_tokens=3000)

def ai_action_plan(summary, domain):
    prompt = f"""You are a senior SEO consultant writing a Prioritized Action Plan for {domain}.

Create a detailed week-by-week and month-by-month action plan. Output pipe-delimited rows:
timeframe|action_item|category|impact|effort|details|owner_notes

Where:
- timeframe = "Week 1", "Week 2", "Month 1", "Month 2-3", "Month 4-6"
- category = "Technical", "On-Page", "Schema", "Content", "Backlinks", "Local SEO"
- impact = "Critical", "High", "Medium", "Low"
- effort = "Low", "Medium", "High"
- details = specific instructions (mention actual URLs, page counts, keywords)
- owner_notes = who should do it and any dependencies

Rules:
- Week 1-2: Critical technical fixes, schema, quick wins
- Month 1: Content launch, on-page optimization, initial backlink work
- Month 2-3: Content scaling, link building ramp-up
- Month 4-6: Authority building, advanced content
- Reference actual data: page counts, keyword volumes, specific URLs
- 25-40 total action items
- No headers or explanations, just the pipe-delimited rows

{summary}

OUTPUT:"""
    return ask_ollama(prompt, max_tokens=5000)

# ─── PARSE AI OUTPUT ─────────────────────────────────────────────────────────
def parse_pipe_rows(text, expected_cols):
    rows = []
    for line in text.strip().split("\n"):
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("---") or line.startswith("SECTION"): continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= expected_cols - 1:
            while len(parts) < expected_cols: parts.append("")
            rows.append(parts[:expected_cols])
    return rows

# ─── EXCEL REPORT ────────────────────────────────────────────────────────────
def generate_report(domain, pages_data, broken_links, ahrefs_data, ai_results, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    BD="1B2A4A"; BA="2E86AB"; BG="2D936C"; BR="C1292E"; BO="E8871E"
    hfont=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    hfill=PatternFill("solid",fgColor=BD)
    secfont=Font(name="Arial",bold=True,size=12,color=BD)
    secfill=PatternFill("solid",fgColor="E8F4F8")
    subfont=Font(name="Arial",size=9,color="666666")
    df=Font(name="Arial",size=9,color="1B1B1B")
    dfb=Font(name="Arial",bold=True,size=9,color="1B1B1B")
    ct=Alignment(horizontal="center",vertical="top")
    lt=Alignment(horizontal="left",vertical="top",wrap_text=True)
    tb=Border(bottom=Side(style="thin",color="DDDDDD"))
    sev_styles={"Critical":(PatternFill("solid",fgColor="FFCDD2"),Font(name="Arial",bold=True,size=9,color=BR)),
                "High":(PatternFill("solid",fgColor="FFE0B2"),Font(name="Arial",bold=True,size=9,color=BO)),
                "Medium":(PatternFill("solid",fgColor="FFF9C4"),Font(name="Arial",bold=True,size=9,color="9E8600")),
                "Low":(PatternFill("solid",fgColor="E8F5E9"),Font(name="Arial",bold=True,size=9,color=BG))}

    def hdr(ws,row,headers):
        for c,h in enumerate(headers,1):
            cell=ws.cell(row=row,column=c,value=h); cell.font=hfont; cell.fill=hfill; cell.alignment=ct
    def colw(ws,widths):
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    def write_row(ws,row,values,bold_first=False):
        for c,v in enumerate(values,1):
            cell=ws.cell(row=row,column=c,value=v)
            cell.font=dfb if (bold_first and c==1) else df
            cell.alignment=lt; cell.border=tb

    dr=ahrefs_data.get("domain_rating",{}); met=ahrefs_data.get("metrics",{}); bls=ahrefs_data.get("backlinks_stats",{})

    # ═══ TAB 1: TECHNICAL AUDIT ═══
    ws=wb.active; ws.title="Technical Audit"; ws.sheet_properties.tabColor=BD
    ws.merge_cells("A1:G1"); ws["A1"]="TECHNICAL SEO AUDIT"; ws["A1"].font=Font(name="Arial",bold=True,size=14,color=BD)
    ws["A2"]=domain; ws["A2"].font=dfb
    info=f"DR: {dr.get('domain_rating','N/A')}  |  Traffic: {met.get('org_traffic',0):,}/mo  |  Keywords: {met.get('org_keywords',0):,}  |  Backlinks: {bls.get('live',0):,}  |  Ref Domains: {bls.get('live_refdomains',0):,}"
    ws["B2"]=info; ws["B2"].font=subfont
    ws["E2"]=f"Audit {datetime.now().strftime('%B %Y')}"; ws["E2"].font=subfont

    r=4; ws.merge_cells(f"A{r}:G{r}"); ws[f"A{r}"]="CRAWL ISSUES"; ws[f"A{r}"].font=secfont; ws[f"A{r}"].fill=secfill
    r+=1; hdr(ws,r,["Issue Category","Severity","Count","Affected Page(s)","Current State","Recommendation","Impact"]); r+=1

    grouped=group_issues(pages_data)
    severity_order={"Critical":0,"High":1,"Medium":2,"Low":3}
    sorted_issues=sorted(grouped.items(), key=lambda x: severity_order.get(ISSUE_META.get(x[0],("Medium",))[0],2))

    for iss_name,iss_data in sorted_issues:
        meta=ISSUE_META.get(iss_name,("Medium","Review and fix.","May affect SEO."))
        severity,recommendation,impact=meta
        page_lines=[]
        for url in iss_data["pages"][:15]:
            detail=iss_data["details"].get(url,""); path=urlparse(url).path or "/"
            page_lines.append(f"{path} ({detail})" if detail else path)
        affected="\n".join(page_lines)
        if len(iss_data["pages"])>15: affected+=f"\n+ {len(iss_data['pages'])-15} more"
        current_state=f"{len(iss_data['pages'])} pages affected across the site."

        ws.cell(row=r,column=1,value=iss_name).font=dfb; ws.cell(row=r,column=1).alignment=lt
        sc=ws.cell(row=r,column=2,value=severity); sc.alignment=ct
        sf,sfont=sev_styles.get(severity,sev_styles["Medium"]); sc.fill=sf; sc.font=sfont
        ws.cell(row=r,column=3,value=f"{len(iss_data['pages'])} pages").font=df; ws.cell(row=r,column=3).alignment=ct
        ws.cell(row=r,column=4,value=affected).font=df; ws.cell(row=r,column=4).alignment=lt
        ws.cell(row=r,column=5,value=current_state).font=df; ws.cell(row=r,column=5).alignment=lt
        ws.cell(row=r,column=6,value=recommendation).font=df; ws.cell(row=r,column=6).alignment=lt
        ws.cell(row=r,column=7,value=impact).font=df; ws.cell(row=r,column=7).alignment=lt
        ws.row_dimensions[r].height=max(30,min(len(page_lines)*13,250))
        for c in range(1,8): ws.cell(row=r,column=c).border=tb
        r+=1

    if broken_links:
        r+=1; ws.merge_cells(f"A{r}:G{r}"); ws[f"A{r}"]="BROKEN LINKS"; ws[f"A{r}"].font=secfont; ws[f"A{r}"].fill=PatternFill("solid",fgColor="FFEBEE"); r+=1
        bl_text="\n".join(f"{urlparse(bl['url']).path} -> {bl['status']}" for bl in broken_links[:20])
        if len(broken_links)>20: bl_text+=f"\n+ {len(broken_links)-20} more"
        write_row(ws,r,["Broken/Error Links","Critical",f"{len(broken_links)} URLs",bl_text,f"{len(broken_links)} URLs return error status codes.","Fix or 301 redirect all broken URLs.","Broken links waste crawl budget."],True)
        sf,sfont=sev_styles["Critical"]; ws.cell(row=r,column=2).fill=sf; ws.cell(row=r,column=2).font=sfont; ws.cell(row=r,column=2).alignment=ct

    colw(ws,[28,10,12,50,45,45,35]); ws.freeze_panes="A6"

    # ═══ TAB 2: CONTENT GAP ═══
    ws2=wb.create_sheet("Content Gap"); ws2.sheet_properties.tabColor=BG
    ws2.merge_cells("A1:G1"); ws2["A1"]="CONTENT GAP ANALYSIS"; ws2["A1"].font=Font(name="Arial",bold=True,size=14,color=BD)
    ws2["A2"]=f"Keywords that {domain} should be targeting"; ws2["A2"].font=subfont
    r=4
    cg_data=ai_results.get("content_gap","")
    if cg_data:
        current_section=""
        rows=parse_pipe_rows(cg_data,7)
        for row in rows:
            section=row[0].strip().upper() if row[0] else ""
            if "PRODUCT" in section or "SERVICE" in section:
                if current_section!="PRODUCT":
                    ws2.merge_cells(f"A{r}:G{r}"); ws2[f"A{r}"]="PRODUCT & SERVICE KEYWORDS (existing pages can target)"; ws2[f"A{r}"].font=secfont; ws2[f"A{r}"].fill=secfill; r+=1
                    hdr(ws2,r,["Keyword","Volume","KD","CPC","Currently Ranking?","Page to Target","Recommendation"]); r+=1
                    current_section="PRODUCT"
                write_row(ws2,r,row[1:8],True); r+=1
            elif "CONTENT" in section or "HIGH" in section or "BLOG" in section:
                if current_section!="CONTENT":
                    r+=1; ws2.merge_cells(f"A{r}:G{r}"); ws2[f"A{r}"]="HIGH-VOLUME CONTENT KEYWORDS (need new blog posts)"; ws2[f"A{r}"].font=secfont; ws2[f"A{r}"].fill=secfill; r+=1
                    hdr(ws2,r,["Keyword","Volume","KD","CPC","Currently Ranking?","Content Type","Blog Title Recommendation"]); r+=1
                    current_section="CONTENT"
                write_row(ws2,r,row[1:8],True); r+=1
            else:
                if current_section=="":
                    ws2.merge_cells(f"A{r}:G{r}"); ws2[f"A{r}"]="KEYWORD OPPORTUNITIES"; ws2[f"A{r}"].font=secfont; ws2[f"A{r}"].fill=secfill; r+=1
                    hdr(ws2,r,["Keyword","Volume","KD","CPC","Currently Ranking?","Page to Target","Recommendation"]); r+=1
                    current_section="GENERAL"
                write_row(ws2,r,row[1:8] if len(row)>7 else row,True); r+=1
    else:
        ws2.cell(row=r,column=1,value="AI analysis was not available. Run with Ollama active.").font=Font(name="Arial",italic=True,color="999999")
    colw(ws2,[25,10,6,10,18,30,50])

    # ═══ TAB 3: BLOG CONTENT STRATEGY ═══
    ws3=wb.create_sheet("Blog Content Strategy"); ws3.sheet_properties.tabColor="6C3483"
    ws3.merge_cells("A1:H1"); ws3["A1"]="BLOG CONTENT STRATEGY"; ws3["A1"].font=Font(name="Arial",bold=True,size=14,color=BD)
    ws3["A2"]=f"Recommended blog posts for {domain}"; ws3["A2"].font=subfont
    r=4; hdr(ws3,r,["Priority","Blog Title","Target Keyword","Volume","KD","Word Count","Internal Link To","Content Notes"]); r+=1
    blog_data=ai_results.get("blog_strategy","")
    if blog_data:
        rows=parse_pipe_rows(blog_data,8)
        for row in rows:
            write_row(ws3,r,row,True)
            pri=row[0].strip().upper()
            sf,sfont=sev_styles.get({"P1":"Critical","P2":"High","P3":"Medium"}.get(pri,"Medium"),sev_styles["Medium"])
            ws3.cell(row=r,column=1).fill=sf; ws3.cell(row=r,column=1).font=sfont; ws3.cell(row=r,column=1).alignment=ct
            r+=1
    colw(ws3,[8,45,22,10,6,12,40,45])

    # ═══ TAB 4: BACKLINK STRATEGY ═══
    ws4=wb.create_sheet("Backlink Strategy"); ws4.sheet_properties.tabColor=BO
    ws4.merge_cells("A1:H1"); ws4["A1"]="BACKLINK PROFILE & LINK BUILDING STRATEGY"; ws4["A1"].font=Font(name="Arial",bold=True,size=14,color=BD)
    ws4["A2"]=f"{domain}  |  {bls.get('live',0):,} live backlinks, {bls.get('live_refdomains',0):,} referring domains"; ws4["A2"].font=subfont
    r=4
    bl_data=ai_results.get("backlink_strategy","")
    if bl_data:
        parts=bl_data.split("---SECTION2---")
        # Section 1: Current profile
        ws4.merge_cells(f"A{r}:H{r}"); ws4[f"A{r}"]="CURRENT BACKLINK PROFILE SNAPSHOT"; ws4[f"A{r}"].font=secfont; ws4[f"A{r}"].fill=secfill; r+=1
        hdr(ws4,r,["Referring Domain","DR","Links","Type","Quality","Action","Notes",""]); r+=1
        profile_rows=parse_pipe_rows(parts[0] if parts else "",7)
        for row in profile_rows:
            write_row(ws4,r,row,True)
            qual=row[4].strip().lower() if len(row)>4 else ""
            if "toxic" in qual or "disavow" in qual:
                ws4.cell(row=r,column=5).fill=PatternFill("solid",fgColor="FFCDD2")
            elif "needs" in qual or "audit" in qual:
                ws4.cell(row=r,column=5).fill=PatternFill("solid",fgColor="FFF9C4")
            elif "legit" in qual:
                ws4.cell(row=r,column=5).fill=PatternFill("solid",fgColor="E8F5E9")
            r+=1
        # Section 2: Link building tactics
        if len(parts)>1:
            r+=1; ws4.merge_cells(f"A{r}:H{r}"); ws4[f"A{r}"]="LINK BUILDING CAMPAIGN"; ws4[f"A{r}"].font=secfont; ws4[f"A{r}"].fill=secfill; r+=1
            hdr(ws4,r,["Tactic","Priority","DR Range","Effort","Approach & Targets","Expected Outcome","Timeline",""]); r+=1
            tactic_rows=parse_pipe_rows(parts[1],7)
            for row in tactic_rows:
                write_row(ws4,r,row,True); r+=1
    else:
        # Fallback: just list referring domains from Ahrefs
        hdr(ws4,r,["Domain","DR","Backlinks","First Seen","Last Visited","","",""]); r+=1
        for rd in ahrefs_data.get("referring_domains",[]):
            write_row(ws4,r,[rd.get("domain",""),rd.get("domain_rating"),rd.get("backlinks"),str(rd.get("first_seen",""))[:10],str(rd.get("last_visited",""))[:10],"","",""]); r+=1
    colw(ws4,[25,8,8,28,14,40,35,25])

    # ═══ TAB 5: ACTION PLAN ═══
    ws5=wb.create_sheet("Action Plan"); ws5.sheet_properties.tabColor=BG
    ws5.merge_cells("A1:G1"); ws5["A1"]="PRIORITIZED SEO ACTION PLAN"; ws5["A1"].font=Font(name="Arial",bold=True,size=14,color=BD)
    ws5["A2"]=f"{domain}  |  {datetime.now().strftime('%B %Y')} start"; ws5["A2"].font=subfont
    r=4; current_tf=""
    ap_data=ai_results.get("action_plan","")
    if ap_data:
        rows=parse_pipe_rows(ap_data,7)
        for row in rows:
            tf=row[0].strip()
            if tf!=current_tf:
                if current_tf: r+=1
                ws5.merge_cells(f"A{r}:G{r}"); ws5[f"A{r}"]=tf.upper(); ws5[f"A{r}"].font=secfont; ws5[f"A{r}"].fill=secfill; r+=1
                hdr(ws5,r,["Action Item","Category","Impact","Effort","Details","Owner/Notes","Timeline"]); r+=1
                current_tf=tf
            write_row(ws5,r,row[1:8],True)
            impact=row[3].strip() if len(row)>3 else ""
            sf,sfont=sev_styles.get(impact,sev_styles.get("Medium"))
            ws5.cell(row=r,column=3).fill=sf; ws5.cell(row=r,column=3).font=sfont; ws5.cell(row=r,column=3).alignment=ct
            r+=1
    colw(ws5,[40,12,10,8,50,30,12])

    wb.save(output_path)

# ─── GUI ─────────────────────────────────────────────────────────────────────
class PassieonApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Passieon SEO Auditor v3.0"); self.geometry("720x700"); self.resizable(False,False)
        ctk.set_appearance_mode("dark"); ctk.set_default_color_theme("blue")
        self.config_data=load_config(); self.running=False; self.build_ui()

    def build_ui(self):
        header=ctk.CTkFrame(self,height=80,fg_color="#1B2A4A",corner_radius=0)
        header.pack(fill="x"); header.pack_propagate(False)
        ctk.CTkLabel(header,text="PASSIEON SEO AUDITOR",font=("Arial",24,"bold"),text_color="white").pack(pady=(15,0))
        ctk.CTkLabel(header,text="Full Crawl  •  Ahrefs  •  AI Strategic Analysis  •  Client Reports",font=("Arial",11),text_color="#8EBBCF").pack()

        main=ctk.CTkFrame(self,fg_color="transparent"); main.pack(fill="both",expand=True,padx=25,pady=15)

        ctk.CTkLabel(main,text="Website URL",font=("Arial",13,"bold")).pack(anchor="w")
        self.url_entry=ctk.CTkEntry(main,placeholder_text="example.com",height=40,font=("Arial",13))
        self.url_entry.pack(fill="x",pady=(3,12))

        kf=ctk.CTkFrame(main,fg_color="transparent"); kf.pack(fill="x")
        ctk.CTkLabel(kf,text="Ahrefs API Key",font=("Arial",13,"bold")).pack(side="left")
        self.key_saved_label=ctk.CTkLabel(kf,text="",font=("Arial",11),text_color="#2D936C"); self.key_saved_label.pack(side="right")
        self.key_entry=ctk.CTkEntry(main,placeholder_text="Enter your Ahrefs API key",height=40,font=("Arial",13),show="*")
        self.key_entry.pack(fill="x",pady=(3,12))
        saved_key=self.config_data.get("ahrefs_api_key","")
        if saved_key: self.key_entry.insert(0,saved_key); self.key_saved_label.configure(text="Saved")

        ctk.CTkLabel(main,text="Target Country",font=("Arial",13,"bold")).pack(anchor="w")
        self.country_menu=ctk.CTkComboBox(main,values=COUNTRIES,height=36,font=("Arial",12),state="readonly")
        self.country_menu.set("Global (all countries)"); self.country_menu.pack(fill="x",pady=(3,15))

        self.ollama_frame=ctk.CTkFrame(main,fg_color="#1a2332",corner_radius=8,height=36)
        self.ollama_frame.pack(fill="x",pady=(0,12))
        self.ollama_label=ctk.CTkLabel(self.ollama_frame,text="",font=("Arial",11)); self.ollama_label.pack(pady=8)
        if check_ollama(): self.ollama_label.configure(text=f"Ollama ready — {OLLAMA_MODEL}",text_color="#2D936C")
        else: self.ollama_label.configure(text=f"Ollama: {OLLAMA_MODEL} not found — run: ollama pull {OLLAMA_MODEL}",text_color="#E8871E")

        bf=ctk.CTkFrame(main,fg_color="transparent"); bf.pack(fill="x",pady=(0,10))
        self.start_btn=ctk.CTkButton(bf,text="START AUDIT",height=44,font=("Arial",14,"bold"),fg_color="#2D936C",hover_color="#24774F",command=self.start_audit)
        self.start_btn.pack(side="left",expand=True,fill="x",padx=(0,5))
        self.open_btn=ctk.CTkButton(bf,text="Open Report",height=44,font=("Arial",13),fg_color="#2E86AB",hover_color="#236B8A",command=self.open_report,state="disabled")
        self.open_btn.pack(side="right",expand=True,fill="x",padx=(5,0))

        self.progress=ctk.CTkProgressBar(main,height=6,progress_color="#2D936C")
        self.progress.pack(fill="x",pady=(0,8)); self.progress.set(0)

        pf=ctk.CTkFrame(main,fg_color="transparent"); pf.pack(fill="x",pady=(0,8))
        self.phase_labels={}
        for ph in ["Crawl","Ahrefs","AI Strategy","Report"]:
            lbl=ctk.CTkLabel(pf,text=ph,font=("Arial",11),text_color="#555555")
            lbl.pack(side="left",expand=True); self.phase_labels[ph]=lbl

        self.log_box=ctk.CTkTextbox(main,height=160,font=("Consolas",11),fg_color="#0d1117",text_color="#8b949e",corner_radius=8)
        self.log_box.pack(fill="both",expand=True)
        self.output_path=None

    def log(self,msg):
        self.log_box.insert("end",msg+"\n"); self.log_box.see("end")
    def set_phase(self,name,status):
        lbl=self.phase_labels.get(name)
        if not lbl: return
        if status=="active": lbl.configure(text_color="#2E86AB",font=("Arial",11,"bold"))
        elif status=="done": lbl.configure(text_color="#2D936C",font=("Arial",11,"bold"))
        elif status=="skip": lbl.configure(text_color="#E8871E")

    def start_audit(self):
        url=self.url_entry.get().strip(); api_key=self.key_entry.get().strip()
        if not url: self.log("Enter a URL."); return
        if not api_key: self.log("Enter API key."); return
        if self.running: return
        self.config_data["ahrefs_api_key"]=api_key; save_config(self.config_data); self.key_saved_label.configure(text="Saved")
        country_sel=self.country_menu.get()
        country=country_sel[:2] if country_sel!="Global (all countries)" else None
        self.running=True; self.start_btn.configure(state="disabled",text="Running...")
        self.open_btn.configure(state="disabled"); self.log_box.delete("1.0","end"); self.progress.set(0)
        for lbl in self.phase_labels.values(): lbl.configure(text_color="#555555",font=("Arial",11))
        threading.Thread(target=self.run_audit,args=(url,api_key,country),daemon=True).start()

    def run_audit(self,url,api_key,country):
        try:
            domain=urlparse(normalize_url(url)).netloc

            # CRAWL (unlimited)
            self.set_phase("Crawl","active"); self.log(f"Crawling {domain} (full site, no limit)...")
            start_url=normalize_url(url); session=requests.Session()
            visited,to_visit=set(),{start_url}; pages_data,broken_links=[],[]
            while to_visit:
                u=to_visit.pop()
                if u in visited: continue
                visited.add(u)
                count=len(visited)
                if count%10==0 or count<10: self.log(f"  [{count}] {urlparse(u).path or '/'}")
                try: resp=session.get(u,headers=HEADERS,timeout=REQUEST_TIMEOUT,allow_redirects=True)
                except: broken_links.append({"url":u,"status":"Error"}); continue
                if resp.status_code>=400: broken_links.append({"url":u,"status":resp.status_code})
                if "text/html" not in resp.headers.get("Content-Type",""): continue
                soup=BeautifulSoup(resp.text,"html.parser")
                pages_data.append(analyze_page(u,resp,soup))
                to_visit.update(extract_links(soup,u,domain)-visited)
                time.sleep(CRAWL_DELAY)
            self.log(f"  Crawl complete: {len(pages_data)} pages, {len(broken_links)} broken links")
            self.set_phase("Crawl","done"); self.progress.set(0.20)

            # AHREFS
            self.set_phase("Ahrefs","active"); self.log("Pulling Ahrefs data...")
            ah=AhrefsAPI(api_key)
            self.log("  Domain Rating..."); d_r=ah.domain_rating(domain)
            self.log("  Metrics..."); met=ah.overview_metrics(domain,country)
            self.log("  Backlinks..."); bstat=ah.backlinks_stats(domain)
            self.log("  Keywords (200)..."); kws=ah.organic_keywords(domain,country,200)
            self.log("  Top Pages (100)..."); tpg=ah.top_pages(domain,100)
            self.log("  Ref Domains (100)..."); rds=ah.referring_domains(domain,100)
            self.log("  Competitors..."); comps=ah.organic_competitors(domain,10)
            ahrefs_data={"domain_rating":d_r,"metrics":met,"backlinks_stats":bstat,
                         "organic_keywords":kws,"top_pages":tpg,"referring_domains":rds,"competitors":comps}
            kw_count=len(kws); pg_count=len(tpg); rd_count=len(rds)
            self.log(f"  Done: {kw_count} keywords, {pg_count} pages, {rd_count} ref domains")
            if kw_count==0: self.log("  [!] No keywords returned — check API key permissions")
            self.set_phase("Ahrefs","done"); self.progress.set(0.35)

            # AI STRATEGY
            ai_results={}
            if check_ollama():
                self.set_phase("AI Strategy","active")
                summary=build_full_summary(domain,pages_data,ahrefs_data)
                country_name=self.country_menu.get()

                self.log("AI: Generating Content Gap Analysis... (1-2 min)")
                ai_results["content_gap"]=ai_content_gap(summary,domain,country_name)
                self.progress.set(0.50)

                self.log("AI: Generating Blog Content Strategy... (1-2 min)")
                ai_results["blog_strategy"]=ai_blog_strategy(summary,domain)
                self.progress.set(0.60)

                self.log("AI: Generating Backlink Strategy... (2-3 min)")
                ai_results["backlink_strategy"]=ai_backlink_strategy(summary,domain)
                self.progress.set(0.75)

                self.log("AI: Generating Action Plan... (2-3 min)")
                ai_results["action_plan"]=ai_action_plan(summary,domain)
                self.progress.set(0.85)

                self.log("  AI analysis complete"); self.set_phase("AI Strategy","done")
            else:
                self.log("Ollama not available — strategic tabs will have data only"); self.set_phase("AI Strategy","skip")
            self.progress.set(0.90)

            # REPORT
            self.set_phase("Report","active"); self.log("Generating Excel report...")
            ts=datetime.now().strftime("%Y%m%d_%H%M"); sd=domain.replace(".","_").replace("www_","")
            if getattr(sys,'frozen',False): out_dir=os.path.dirname(sys.executable)
            else: out_dir=os.path.dirname(os.path.abspath(__file__))
            self.output_path=os.path.join(out_dir,f"Passieon_SEO_Audit_{sd}_{ts}.xlsx")
            generate_report(domain,pages_data,broken_links,ahrefs_data,ai_results,self.output_path)
            self.log(f"  Saved: {os.path.basename(self.output_path)}")
            self.set_phase("Report","done"); self.progress.set(1.0)
            self.log(f"\nAUDIT COMPLETE — {len(pages_data)} pages crawled")
            self.start_btn.configure(state="normal",text="START AUDIT"); self.open_btn.configure(state="normal"); self.running=False
        except Exception as e:
            self.log(f"\nError: {e}"); import traceback; self.log(traceback.format_exc())
            self.start_btn.configure(state="normal",text="START AUDIT"); self.running=False

    def open_report(self):
        if self.output_path and os.path.exists(self.output_path):
            os.startfile(self.output_path) if sys.platform=="win32" else os.system(f"open '{self.output_path}'")

if __name__=="__main__":
    app=PassieonApp(); app.mainloop()
